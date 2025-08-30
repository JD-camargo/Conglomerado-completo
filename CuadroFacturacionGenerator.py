# CuadroFacturacionGenerator.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class CuadroFacturacionGenerator:
    def __init__(self, filename="Cuadro_Facturacion.xlsx"):
        self.filename = filename

    def generar(self, datos):
        """
        Genera un cuadro de facturación en Excel.

        datos: lista de diccionarios con claves:
            - 'item'
            - 'descripcion'
            - 'cantidad'
            - 'valor_unitario'
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturación"

        # Encabezados
        headers = ["Item", "Descripción", "Cantidad", "Valor Unitario", "Subtotal"]
        ws.append(headers)

        # Estilo encabezados
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Insertar datos
        for i, dato in enumerate(datos, start=2):
            ws.cell(row=i, column=1, value=dato["item"])
            ws.cell(row=i, column=2, value=dato["descripcion"])
            ws.cell(row=i, column=3, value=dato["cantidad"])
            ws.cell(row=i, column=4, value=dato["valor_unitario"])
            ws.cell(row=i, column=5, value=dato["cantidad"] * dato["valor_unitario"])

        # Total
        total_row = len(datos) + 2
        ws.cell(row=total_row, column=4, value="TOTAL")
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{len(datos)+1})")

        # Bordes
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=1, max_row=total_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        # Guardar archivo
        wb.save(self.filename)
        return self.filename
